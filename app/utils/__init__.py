import hashlib
import json
from datetime import datetime
from typing import Dict, Any, Optional
from fastapi import UploadFile, HTTPException
from app.core.config import settings


def normalize_slug(name: str) -> str:
    import re
    normalized = re.sub(r"[^a-zA-Z0-9\s]", "", name.lower())
    normalized = re.sub(r"\s+", "-", normalized.strip())
    return normalized[:100]


def secure_filename(filename: str) -> str:
    import re
    filename = re.sub(r"[^a-zA-Z0-9\.\-]", "_", filename)
    return filename[:255]


def is_allowed_extension(filename: str, allowed_extensions: list) -> bool:
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions


def generate_certificate_id(prefix: str, year: int, sequence: int) -> str:
    return f"{prefix}-{year}-{sequence:06d}"


def generate_qr_code(data: Dict[str, Any]) -> str:
    import hashlib
    import json
    
    json_data = json.dumps(data, separators=(',', ':'))
    token = hashlib.sha256(json_data.encode()).hexdigest()[:64]
    return token


def generate_watermark(text: str, opacity: float = 0.3) -> Any:
    from PIL import Image, ImageDraw, ImageFont
    import io
    
    img = Image.new('RGBA', (100, 100), (255, 255, 255, 0))
    draw = ImageDraw.Draw(img)
    
    try:
        font = ImageFont.truetype("arial.ttf", 20)
    except Exception:
        font = ImageFont.load_default()
    
    text_bbox = draw.textbbox((0, 0), text, font=font)
    text_width = text_bbox[2] - text_bbox[0]
    text_height = text_bbox[3] - text_bbox[1]
    
    x = (img.width - text_width) // 2
    y = (img.height - text_height) // 2
    
    draw.text((x, y), text, font=font, fill=(128, 128, 128, int(opacity * 255)))
    
    return img


def compute_tamper_hash(participant: Any, certificate_id: str, event: str, date: str, secret: str) -> str:
    from pydantic import BaseModel

    data = {
        "participant_email": participant.email,
        "participant_name": participant.name,
        "participant_event": event,
        "certificate_id": certificate_id,
        "date": date
    }
    
    data_string = json.dumps(data, sort_keys=True)
    
    hash_input = data_string + secret
    
    hash_obj = hashlib.sha256()
    hash_obj.update(hash_input.encode('utf-8'))
    
    return hash_obj.hexdigest()


def verify_tamper_hash(participant: Any, certificate_id: str, event: str, date: str, stored_hash: str, secret: str) -> bool:
    computed_hash = compute_tamper_hash(participant, certificate_id, event, date, secret)
    return computed_hash == stored_hash


def generate_unique_filename(file_type: str, base_name: str) -> str:
    import os
    from datetime import datetime
    now = datetime.utcnow()
    timestamp = now.strftime("%Y%m%d%H%M%S")
    storage_base = settings.STORAGE_BASE_PATH
    return os.path.join(storage_base, "generated", file_type, f"{timestamp}_{base_name}.{file_type}")


def validate_participant_data(data: Dict[str, Any], organization_id: str) -> Dict[str, Any]:
    errors = []
    
    required_fields = ["name", "email"]
    for field in required_fields:
        if not data.get(field):
            errors.append(f"{field} is required")
    
    email = data.get("email")
    if email and "@" not in email:
        errors.append("Invalid email format")
    
    return {"data": data, "errors": errors} if errors else {"data": data, "errors": []}


def parse_csv_file(file_content: bytes, field_mapping: Optional[dict] = None) -> list:
    import csv
    from io import StringIO
    
    text_content = file_content.decode('utf-8-sig')
    participants = []
    
    reader = csv.DictReader(StringIO(text_content))
    
    raw_headers = [h.strip() for h in (reader.fieldnames or [])]
    normalized_headers = {h.lower(): h for h in raw_headers}
    
    for row in reader:
        participant = {}
        if field_mapping:
            for original_field, mapped_field in field_mapping.items():
                key = original_field.strip().lower()
                actual_header = normalized_headers.get(key)
                if actual_header and actual_header in row:
                    val = row[actual_header]
                    participant[mapped_field] = val.strip() if isinstance(val, str) else val

        if not participant and isinstance(row, dict):
            for k, v in row.items():
                if v and k.strip():
                    normal_key = k.strip().lower()
                    val = v.strip() if isinstance(v, str) else v
                    participant[normal_key] = val

        if participant:
            participants.append(participant)
    
    return participants


def parse_excel_file(file_content: bytes, field_mapping: Optional[dict] = None) -> list:
    import io
    from openpyxl import load_workbook

    file_like = io.BytesIO(file_content)
    workbook = load_workbook(file_like, data_only=True)
    worksheet = workbook.active

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip().lower() for h in rows[0]]

    normalized_mapping = {k.strip().lower(): v for k, v in (field_mapping or {}).items()}

    participants = []
    for row in rows[1:]:
        if not any(row):
            continue

        data = {}
        for index, value in enumerate(row):
            if index >= len(headers):
                continue
            excel_header = headers[index]
            if excel_header in normalized_mapping:
                field = normalized_mapping[excel_header]
                data[field] = str(value).strip() if isinstance(value, str) else value

        if not data:
            data = {}
            for index, value in enumerate(row):
                if index >= len(headers):
                    continue
                key = headers[index].strip().lower()
                if value is not None:
                    data[key] = str(value).strip() if isinstance(value, str) else value

        if data:
            participants.append(data)

    return participants


def get_file_size_mb(file: UploadFile) -> float:
    file.file.seek(0, 2)
    size = file.file.tell()
    file.file.seek(0)
    return size / (1024 * 1024)